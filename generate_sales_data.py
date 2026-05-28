from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PRODUCT_SHEETS = {
    "MS": {"sales_unit_field": "Sales Unit", "net_volume_field": "Volume"},
    "HSD": {"sales_unit_field": "Sales Unit", "net_volume_field": "Quantity"},
    "Power": {"sales_unit_field": "Sales Unit", "net_volume_field": "Quantity"},
    "Lube": {"sales_unit_field": "Volume unit", "net_volume_field": "Volume"},
    "DEF": {"sales_unit_field": "Volume unit", "net_volume_field": "Volume"},
}

BASE_PRODUCTS = list(PRODUCT_SHEETS.keys())

RAW_EXPORT_ALIASES = {
    "sapCode": ("shiptoparty", "sapcode", "outletcode", "dealercode"),
    "outletName": ("shiptopartyname", "outletname", "dealername", "nameofdealer"),
    "material": ("material", "materialcode"),
    "materialDescription": ("materialdescription", "itemdescription", "materialdesc", "description"),
    "volume": ("volume", "quantity", "netvolume", "netquantity"),
    "unit": ("salesunit", "volumeunit"),
    "date": ("billingdate", "date"),
    "billingDocument": ("billingdocumentno", "billingdocument", "documentno", "invoiceno", "invoice"),
    "plant": ("plant",),
    "plantDescription": ("plantdescription",),
    "salesArea": ("salesgroupdesc", "salesarea", "salesgroup", "salesareafull"),
    "outletType": ("outlettype", "projectabhuyaday", "paoutlet"),
    "documents": ("documents", "documentcount"),
}

OUTLET_MASTER_ALIASES = {
    "sapCode": ("sapcode", "shiptoparty", "outletcode", "dealercode"),
    "outletName": ("nameofdealership", "shiptopartyname", "outletname", "dealername"),
    "location": ("location",),
    "district": ("district",),
    "salesArea": ("slsarea", "salesarea", "salesgroupdesc", "salesgroup"),
    "depot": ("depot", "plant", "plantdescription"),
    "appointedOn": ("dealershipagreementdate", "dateofcomm", "dateofcommissioning"),
    "agreementExpiry": ("dealershipagreementexpirydate", "agreementexpirydate"),
    "projectedMs": ("projectedavgvolasperdaforms",),
    "projectedHsd": ("projectedavgvolasperdaforhsd",),
    "aopMs": ("aopsalestarget201920forms",),
    "aopHsd": ("aopsalestarget201920forhsd",),
    "classOfMarket": ("classofmarket",),
    "socialCategory": ("socialcategory",),
    "ownership": ("ownershipcocldo", "ownership"),
}

ABHUYADAY_ALIASES = {
    "sapCode": ("customercode", "sapcode", "shiptoparty", "outletcode", "dealercode"),
    "outletName": ("outletname", "shiptopartyname", "dealername", "nameofdealer"),
    "salesArea": ("salesarea", "salesgroupdesc", "slsarea"),
    "msHistorical": ("ms2526", "mshist", "mshistorical", "mshistoricalkl"),
    "hsdHistorical": ("hsd2526", "hsdhist", "hsdhistorical", "hsdhistoricalkl"),
    "msTarget": ("ms2627", "mstarget", "mstargetkl"),
    "hsdTarget": ("hsd2627", "hsdtarget", "hsdtargetkl"),
}

DEPOT_ALIASES = {
    "1424": "JABALPUR DEPOT",
    "JABALPUR": "JABALPUR DEPOT",
    "JABALPUR DEPOT": "JABALPUR DEPOT",
    "1436": "SAGAR DEPOT",
    "SAGAR": "SAGAR DEPOT",
    "SAGAR DEPOT": "SAGAR DEPOT",
}


def classify_product(default_product: str, material_description: object, material: object) -> str:
    description = str(material_description or "").strip().upper()
    description = re.sub(r"\s+", " ", description)
    material_text = str(material or "").strip().upper()
    signal = f"{default_product} {description} {material_text}".upper()

    # Raw export material mapping: P95 is Power, normal EBMS is MS, and HSD BS VI is HSD.
    if re.search(r"^EBMS\s+P95-HSN-27101242\b", description) or re.search(r"\bP95\b|\bPOWER\b", signal):
        return "Power"
    if re.search(r"^HSD\s*-\s*BS\s*VI\b", description) or re.search(r"\bHSD\b", signal):
        return "HSD"
    if re.search(r"\bLUBE\b|\bLUBRICANT\b|\bLUB\b", signal):
        return "Lube"
    if re.search(r"\bDEF\b|\bAUS32\b|\bUREA\b", signal):
        return "DEF"
    if re.search(r"\bE-20\b|\bE20\b", signal):
        return "E20"
    if re.search(r"^EBMS-HSN-2710124[23]\b", description) or re.search(r"\bMS\b|\bPETROL\b|\bEBMS\b", signal):
        return "MS"
    return default_product


def normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def parse_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_date(value: object) -> str | None:
    if value is None:
        return None
    if hasattr(value, "date"):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return None
    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except ValueError:
        return None


def derive_sales_kl(value: object, unit: object) -> float:
    numeric_value = parse_number(value)
    if not numeric_value:
        return 0.0
    normalized_unit = str(unit or "").strip().upper()
    if normalized_unit in {"L", "LTR", "LITRE", "LITRES"}:
        return numeric_value / 1000
    if normalized_unit == "KL":
        return numeric_value
    return numeric_value / 1000 if numeric_value >= 1000 else numeric_value


def normalize_area(value: object) -> str:
    if value is None:
        return "UNKNOWN"
    text = str(value).strip().upper()
    return text.replace(" RETAIL SA", "")


def normalize_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value))
    return str(value).strip()


def normalize_depot(value: object) -> str:
    text = normalize_code(value).strip().upper()
    text = re.sub(r"\s+", " ", text)
    if not text:
        return ""
    return DEPOT_ALIASES.get(text, text)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate dashboard-friendly sales data from the daily sales workbook."
    )
    parser.add_argument(
        "--input",
        nargs="+",
        default=[r"C:\Users\Rohit\Desktop\try.XLSX"],
        help="Path to one or more Excel workbooks.",
    )
    parser.add_argument(
        "--output",
        default=str(Path(__file__).with_name("sales-data.js")),
        help="Path to the generated JavaScript data file.",
    )
    parser.add_argument(
        "--historical-input",
        default="",
        help="Optional raw export workbook to use for MS/HSD/Power Hist. lookup values.",
    )
    parser.add_argument(
        "--outlet-master",
        default=r"C:\Users\Rohit\Desktop\JRRO Total outlets.xlsx",
        help="Optional outlet master workbook to seed the complete retail outlet list.",
    )
    parser.add_argument(
        "--abhuyaday-input",
        default=r"C:\Users\Rohit\Desktop\abhuvaday.xlsx",
        help="Optional Project Abhuyaday workbook with outlet, historical, and target values.",
    )
    return parser.parse_args()


def read_pa_codes(workbook, fallback_payload: dict | None = None) -> set[str]:
    if "PA Outlets" not in workbook.sheetnames:
        fallback_payload = fallback_payload or {}
        meta_codes = fallback_payload.get("meta", {}).get("projectAbhuyadayCodes", [])
        outlet_codes = [
            outlet.get("sapCode")
            for outlet in fallback_payload.get("filters", {}).get("outlets", [])
            if outlet.get("isProjectAbhuyaday")
        ]
        return {normalize_code(code) for code in [*meta_codes, *outlet_codes] if normalize_code(code)}

    codes: set[str] = set()
    ws = workbook["PA Outlets"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = normalize_code(row[0])
        if code:
            codes.add(code)
    return codes


def parse_project_flag(value: object, sap_code: str, pa_codes: set[str]) -> bool:
    signal = str(value or "").strip().upper()
    if signal in {"YES", "TRUE", "1"} or "PROJECT" in signal or "ABHUYADAY" in signal:
        return True
    if signal in {"NO", "FALSE", "0", "STANDARD"}:
        return False
    return sap_code in pa_codes


def product_sequence_from_rows(rows: list[dict]) -> list[str]:
    extras = sorted({row["product"] for row in rows if row["product"] and row["product"] not in BASE_PRODUCTS})
    return [*BASE_PRODUCTS, *extras]


def build_product_meta(rows: list[dict], product_sequence: list[str]) -> dict[str, dict]:
    product_meta: dict[str, dict] = {}
    for product in product_sequence:
        product_rows = [item for item in rows if item["product"] == product]
        dates = [item["date"] for item in product_rows]
        outlets = {item["sapCode"] for item in product_rows}
        areas = sorted({item["salesArea"] for item in product_rows})
        pa_outlets = {item["sapCode"] for item in product_rows if item["isProjectAbhuyaday"]}

        product_meta[product] = {
            "records": len(product_rows),
            "outlets": len(outlets),
            "areas": areas,
            "dateMin": min(dates) if dates else None,
            "dateMax": max(dates) if dates else None,
            "paOutlets": len(pa_outlets),
            "salesUnits": round(sum(item["salesUnits"] for item in product_rows), 3),
            "netVolume": round(sum(item["netVolume"] for item in product_rows), 3),
        }
    return product_meta


def load_product_sheet_rows(workbook, pa_codes: set[str]) -> list[dict]:
    rows: list[dict] = []

    for product, config in PRODUCT_SHEETS.items():
        ws = workbook[product]
        headers = None

        for raw_row in ws.iter_rows(values_only=True):
            values = list(raw_row)
            if headers is None and any(value is not None for value in values):
                headers = values
                continue
            if headers is None or not any(value is not None for value in values):
                continue

            record = dict(zip(headers, values))
            sap_code = normalize_code(record.get("Ship-to Party"))
            billing_date = record.get("Billing Date")

            if not sap_code or billing_date is None:
                continue

            sales_area_full = str(record.get("Sales Group Desc") or "UNKNOWN").strip().upper()
            sales_area = normalize_area(sales_area_full)
            outlet_name = str(record.get("Ship-to-Party Name") or "UNKNOWN OUTLET").strip()
            sales_units = parse_number(record.get(config["sales_unit_field"]))
            net_volume = parse_number(record.get(config["net_volume_field"]))
            bill_no = normalize_code(record.get("Billing Document No."))
            material = normalize_code(record.get("Material"))
            material_description = str(record.get("Material Description") or "").strip()
            plant = normalize_depot(record.get("Plant"))
            iso_date = normalize_date(billing_date)
            classified_product = classify_product(product, material_description, material)
            if not iso_date or not classified_product:
                continue

            entry = {
                "product": classified_product,
                "date": iso_date,
                "salesArea": sales_area,
                "salesAreaFull": sales_area_full,
                "sapCode": sap_code,
                "outletName": outlet_name,
                "salesUnits": round(sales_units, 3),
                "netVolume": round(net_volume, 3),
                "billingDocument": bill_no,
                "material": material,
                "materialDescription": material_description,
                "plant": plant,
                "isProjectAbhuyaday": sap_code in pa_codes,
            }
            rows.append(entry)

    return rows


def find_header_index(headers: list[object], aliases: tuple[str, ...]) -> int | None:
    normalized_headers = [normalize_header(header) for header in headers]
    for alias in aliases:
        if alias in normalized_headers:
            return normalized_headers.index(alias)
    return None


def get_indexed_value(values: list[object], index: int | None) -> object:
    if index is None or index >= len(values):
        return None
    return values[index]


def load_outlet_master_entries(master_path: Path | None) -> list[dict]:
    if not master_path or not master_path.exists():
        return []

    workbook = load_workbook(master_path, read_only=True, data_only=True)
    try:
        ws = workbook[workbook.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        headers = None
        for row in rows:
            values = list(row)
            if any(value is not None for value in values):
                headers = values
                break

        if not headers:
            return []

        indexes = {key: find_header_index(headers, aliases) for key, aliases in OUTLET_MASTER_ALIASES.items()}
        required = ("sapCode", "outletName", "salesArea")
        if any(indexes[key] is None for key in required):
            missing = ", ".join(key for key in required if indexes[key] is None)
            raise ValueError(f"Outlet master sheet is missing required columns: {missing}")

        entries: list[dict] = []
        seen: set[str] = set()
        for row in rows:
            values = list(row)
            if not any(value is not None for value in values):
                continue

            sap_code = normalize_code(get_indexed_value(values, indexes["sapCode"]))
            outlet_name = str(get_indexed_value(values, indexes["outletName"]) or "").strip()
            sales_area = normalize_area(get_indexed_value(values, indexes["salesArea"]))
            if not sap_code or not outlet_name or not sales_area or sap_code in seen:
                continue

            seen.add(sap_code)
            depot = normalize_depot(get_indexed_value(values, indexes["depot"]))
            appointed_on = normalize_date(
                get_indexed_value(values, indexes["appointedOn"])
                or get_indexed_value(values, find_header_index(headers, ("dateofcomm",)))
            )
            agreement_expiry = normalize_date(get_indexed_value(values, indexes["agreementExpiry"]))
            notes = []
            district = str(get_indexed_value(values, indexes["district"]) or "").strip()
            location = str(get_indexed_value(values, indexes["location"]) or "").strip()
            class_of_market = str(get_indexed_value(values, indexes["classOfMarket"]) or "").strip()
            ownership = str(get_indexed_value(values, indexes["ownership"]) or "").strip()
            social_category = str(get_indexed_value(values, indexes["socialCategory"]) or "").strip()
            if location:
                notes.append(f"Location: {location}")
            if district:
                notes.append(f"District: {district}")
            if class_of_market:
                notes.append(f"Market: {class_of_market}")
            if ownership:
                notes.append(f"Ownership: {ownership}")
            if social_category:
                notes.append(f"Category: {social_category}")
            if agreement_expiry:
                notes.append(f"Agreement expiry: {agreement_expiry}")

            entries.append(
                {
                    "sapCode": sap_code,
                    "outletName": outlet_name,
                    "salesArea": sales_area,
                    "isProjectAbhuyaday": False,
                    "appointedOn": appointed_on,
                    "plant": depot,
                    "notes": " | ".join(notes),
                    "outletType": "Standard",
                    "masterSource": master_path.name,
                    "targets": {
                        "MS": round(parse_number(get_indexed_value(values, indexes["aopMs"])), 3),
                        "HSD": round(parse_number(get_indexed_value(values, indexes["aopHsd"])), 3),
                    },
                }
            )
        return entries
    finally:
        workbook.close()


def load_abhuyaday_entries(abhuyaday_path: Path | None) -> list[dict]:
    if not abhuyaday_path or not abhuyaday_path.exists():
        return []

    workbook = load_workbook(abhuyaday_path, read_only=True, data_only=True)
    try:
        ws = workbook[workbook.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        headers = None
        for row in rows:
            values = list(row)
            if any(value is not None for value in values):
                headers = values
                break

        if not headers:
            return []

        indexes = {key: find_header_index(headers, aliases) for key, aliases in ABHUYADAY_ALIASES.items()}
        required = ("sapCode", "outletName", "salesArea")
        if any(indexes[key] is None for key in required):
            missing = ", ".join(key for key in required if indexes[key] is None)
            raise ValueError(f"Project Abhuyaday sheet is missing required columns: {missing}")

        entries: list[dict] = []
        seen: set[str] = set()
        for row in rows:
            values = list(row)
            if not any(value is not None for value in values):
                continue

            sap_code = normalize_code(get_indexed_value(values, indexes["sapCode"]))
            outlet_name = str(get_indexed_value(values, indexes["outletName"]) or "").strip()
            sales_area = normalize_area(get_indexed_value(values, indexes["salesArea"]))
            if not sap_code or not outlet_name or not sales_area or sap_code in seen:
                continue

            seen.add(sap_code)
            ms_historical = round(parse_number(get_indexed_value(values, indexes["msHistorical"])), 3)
            hsd_historical = round(parse_number(get_indexed_value(values, indexes["hsdHistorical"])), 3)
            ms_target = round(parse_number(get_indexed_value(values, indexes["msTarget"])), 3)
            hsd_target = round(parse_number(get_indexed_value(values, indexes["hsdTarget"])), 3)

            entries.append(
                {
                    "sapCode": sap_code,
                    "outletName": outlet_name,
                    "salesArea": sales_area,
                    "isProjectAbhuyaday": True,
                    "outletType": "Project Abhuyaday",
                    "masterSource": abhuyaday_path.name,
                    "notes": f"Project Abhuyaday source: {abhuyaday_path.name}",
                    "historical": {
                        "MS": ms_historical if ms_historical > 0 else 0,
                        "HSD": hsd_historical if hsd_historical > 0 else 0,
                    },
                    "targets": {
                        "MS": ms_target if ms_target > 0 else 0,
                        "HSD": hsd_target if hsd_target > 0 else 0,
                    },
                }
            )
        return entries
    finally:
        workbook.close()


def merge_abhuyaday_outlets(outlet_master_entries: list[dict], abhuyaday_entries: list[dict]) -> list[dict]:
    if not abhuyaday_entries:
        return outlet_master_entries

    merged = {entry["sapCode"]: dict(entry) for entry in outlet_master_entries}
    for entry in abhuyaday_entries:
        existing = merged.get(entry["sapCode"])
        if existing:
            notes = [note for note in (existing.get("notes"), entry.get("notes")) if note]
            merged[entry["sapCode"]] = {
                **existing,
                "outletName": existing.get("outletName") or entry["outletName"],
                "salesArea": entry.get("salesArea") or existing.get("salesArea", "UNKNOWN"),
                "isProjectAbhuyaday": True,
                "outletType": "Project Abhuyaday",
                "notes": " | ".join(dict.fromkeys(notes)),
                "targets": {
                    **(existing.get("targets") or {}),
                    **(entry.get("targets") or {}),
                },
            }
            continue

        merged[entry["sapCode"]] = {
            "sapCode": entry["sapCode"],
            "outletName": entry["outletName"],
            "salesArea": entry["salesArea"],
            "isProjectAbhuyaday": True,
            "appointedOn": None,
            "plant": "",
            "notes": entry.get("notes", ""),
            "outletType": "Project Abhuyaday",
            "masterSource": entry.get("masterSource", ""),
            "targets": entry.get("targets", {}),
        }

    return sorted(merged.values(), key=lambda item: (item["salesArea"], item["outletName"]))


def build_abhuyaday_area_targets(abhuyaday_entries: list[dict]) -> list[dict]:
    if not abhuyaday_entries:
        return []

    area_totals: dict[str, dict[str, float]] = defaultdict(lambda: {"MS": 0.0, "HSD": 0.0})
    source_file = abhuyaday_entries[0].get("masterSource", "abhuvaday.xlsx")
    for entry in abhuyaday_entries:
        sales_area = entry.get("salesArea") or "UNKNOWN"
        targets = entry.get("targets") or {}
        for product in ("MS", "HSD"):
            target_kl = parse_number(targets.get(product))
            if target_kl > 0:
                area_totals[sales_area][product] += target_kl

    updated_at = datetime.now().isoformat(timespec="seconds")
    area_targets = []
    for sales_area, totals in sorted(area_totals.items()):
        monthly_ms = round(totals["MS"] / 12, 3) if totals["MS"] > 0 else 0
        monthly_hsd = round(totals["HSD"] / 12, 3) if totals["HSD"] > 0 else 0
        if monthly_ms <= 0 and monthly_hsd <= 0:
            continue
        area_targets.append(
            {
                "monthKey": "",
                "salesArea": sales_area,
                "targets": {
                    "MS": monthly_ms,
                    "HSD": monthly_hsd,
                    "Power": 0,
                    "Lube": 0,
                    "DEF": 0,
                },
                "note": f"Project Abhuyaday FY 26-27 target from {source_file}, divided by 12 for monthly comparison.",
                "updatedAt": updated_at,
            }
        )
    return area_targets


def build_abhuyaday_reference(abhuyaday_entries: list[dict]) -> dict:
    if not abhuyaday_entries:
        return {"outlets": [], "areaSummary": []}

    area_summary: dict[str, dict] = defaultdict(
        lambda: {
            "salesArea": "",
            "outlets": 0,
            "msHistorical": 0.0,
            "hsdHistorical": 0.0,
            "msTarget": 0.0,
            "hsdTarget": 0.0,
        }
    )
    reference_outlets = []

    for entry in abhuyaday_entries:
        historical = entry.get("historical") or {}
        targets = entry.get("targets") or {}
        ms_historical = parse_number(historical.get("MS"))
        hsd_historical = parse_number(historical.get("HSD"))
        ms_target = parse_number(targets.get("MS"))
        hsd_target = parse_number(targets.get("HSD"))
        reference_outlets.append(
            {
                "sapCode": entry["sapCode"],
                "outletName": entry["outletName"],
                "salesArea": entry["salesArea"],
                "historical": {
                    "MS": round(ms_historical, 3),
                    "HSD": round(hsd_historical, 3),
                },
                "targets": {
                    "MS": round(ms_target, 3),
                    "HSD": round(hsd_target, 3),
                },
            }
        )

        target = area_summary[entry["salesArea"]]
        target["salesArea"] = entry["salesArea"]
        target["outlets"] += 1
        target["msHistorical"] += ms_historical if ms_historical > 0 else 0
        target["hsdHistorical"] += hsd_historical if hsd_historical > 0 else 0
        target["msTarget"] += ms_target if ms_target > 0 else 0
        target["hsdTarget"] += hsd_target if hsd_target > 0 else 0

    return {
        "outlets": reference_outlets,
        "areaSummary": [
            {
                **row,
                "msHistorical": round(row["msHistorical"], 3),
                "hsdHistorical": round(row["hsdHistorical"], 3),
                "msTarget": round(row["msTarget"], 3),
                "hsdTarget": round(row["hsdTarget"], 3),
            }
            for row in sorted(area_summary.values(), key=lambda item: item["salesArea"])
        ],
    }


def apply_outlet_master_to_rows(rows: list[dict], outlet_master_entries: list[dict]) -> list[dict]:
    if not outlet_master_entries:
        return rows

    master_map = {entry["sapCode"]: entry for entry in outlet_master_entries}
    updated_rows: list[dict] = []
    for row in rows:
        master = master_map.get(row["sapCode"])
        if not master:
            updated_rows.append(row)
            continue
        updated_rows.append(
            {
                **row,
                "outletName": master["outletName"] or row["outletName"],
                "salesArea": master["salesArea"] or row["salesArea"],
                "salesAreaFull": f'{master["salesArea"]} RETAIL SA' if master.get("salesArea") else row.get("salesAreaFull"),
                "plant": normalize_depot(master.get("plant") or row.get("plant", "")),
                "isProjectAbhuyaday": bool(master.get("isProjectAbhuyaday") or row.get("isProjectAbhuyaday", False)),
            }
        )
    return updated_rows


def load_raw_export_rows(workbook, pa_codes: set[str]) -> list[dict]:
    ws = workbook[workbook.sheetnames[0]]
    raw_rows = ws.iter_rows(values_only=True)
    headers = None

    for row in raw_rows:
        values = list(row)
        if any(value is not None for value in values):
            headers = values
            break

    if not headers:
        return []

    indexes = {key: find_header_index(headers, aliases) for key, aliases in RAW_EXPORT_ALIASES.items()}
    required = ("sapCode", "materialDescription", "volume", "date", "salesArea")
    if any(indexes[key] is None for key in required):
        missing = ", ".join(key for key in required if indexes[key] is None)
        raise ValueError(f"Raw export sheet is missing required columns: {missing}")

    rows: list[dict] = []
    for row in raw_rows:
        values = list(row)
        if not any(value is not None for value in values):
            continue

        sap_code = normalize_code(get_indexed_value(values, indexes["sapCode"]))
        iso_date = normalize_date(get_indexed_value(values, indexes["date"]))
        material = normalize_code(get_indexed_value(values, indexes["material"]))
        material_description = str(get_indexed_value(values, indexes["materialDescription"]) or "").strip()
        classified_product = classify_product("", material_description, material)
        if not sap_code or not iso_date or not classified_product:
            continue

        volume = get_indexed_value(values, indexes["volume"])
        unit = get_indexed_value(values, indexes["unit"]) or "L"
        sales_units = derive_sales_kl(volume, unit)
        net_volume = parse_number(volume)
        bill_no = normalize_code(get_indexed_value(values, indexes["billingDocument"]))
        if not sales_units:
            continue

        sales_area_full = str(get_indexed_value(values, indexes["salesArea"]) or "UNKNOWN").strip().upper()
        outlet_name = str(get_indexed_value(values, indexes["outletName"]) or "UNKNOWN OUTLET").strip()
        documents = max(1, round(parse_number(get_indexed_value(values, indexes["documents"])) or (1 if bill_no else 1)))

        rows.append(
            {
                "product": classified_product,
                "date": iso_date,
                "salesArea": normalize_area(sales_area_full),
                "salesAreaFull": sales_area_full,
                "sapCode": sap_code,
                "outletName": outlet_name,
                "salesUnits": round(sales_units, 3),
                "netVolume": round(net_volume, 3),
                "billingDocument": bill_no,
                "material": material,
                "materialDescription": material_description,
                "plant": normalize_depot(
                    get_indexed_value(values, indexes["plantDescription"]) or get_indexed_value(values, indexes["plant"])
                ),
                "isProjectAbhuyaday": parse_project_flag(get_indexed_value(values, indexes["outletType"]), sap_code, pa_codes),
                "documents": documents,
            }
        )

    return rows


def load_rows(workbook, pa_codes: set[str]) -> tuple[list[dict], dict]:
    if all(sheet_name in workbook.sheetnames for sheet_name in PRODUCT_SHEETS):
        rows = load_product_sheet_rows(workbook, pa_codes)
    else:
        rows = load_raw_export_rows(workbook, pa_codes)

    rows.sort(
        key=lambda item: (
            item["date"],
            item["salesArea"],
            item["outletName"],
            item["product"],
            item["billingDocument"],
        )
    )

    product_sequence = product_sequence_from_rows(rows)
    return rows, build_product_meta(rows, product_sequence)


def normalize_month_day(value: object) -> str | None:
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%m-%d")
    text = str(value).strip()
    if not text:
        return None
    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            parsed = datetime.strptime(text, pattern)
            return parsed.strftime("%m-%d")
        except ValueError:
            continue
    try:
        parsed = datetime.fromisoformat(text)
        return parsed.strftime("%m-%d")
    except ValueError:
        return None


def fallback_historical_lookup(payload: dict | None) -> dict:
    payload = payload or {}
    existing = payload.get("historicalLookup", {})
    return {
        "MS": dict(existing.get("MS", {})),
        "HSD": dict(existing.get("HSD", {})),
        "Power": dict(existing.get("Power", {})),
    }


def build_historical_lookup_from_rows(rows: list[dict], fallback_payload: dict | None = None) -> dict:
    lookup = {"MS": {}, "HSD": {}, "Power": {}}
    buckets: dict[str, dict[str, dict[str, float]]] = {"MS": {}, "HSD": {}, "Power": {}}

    for row in rows:
        product = row.get("product")
        if product not in buckets:
            continue
        sap_code = normalize_code(row.get("sapCode"))
        date = str(row.get("date") or "")
        if not sap_code or len(date) < 10:
            continue
        lookup_key = f"{sap_code}|{date[5:]}"
        sales_kl = parse_number(row.get("salesUnits"))
        if sales_kl <= 0:
            continue
        dated_bucket = buckets[product].setdefault(lookup_key, {})
        dated_bucket[date] = round(dated_bucket.get(date, 0.0) + sales_kl, 3)

    for product, product_buckets in buckets.items():
        for lookup_key, dated_bucket in product_buckets.items():
            first_historical_date = sorted(dated_bucket)[0]
            lookup[product][lookup_key] = round(dated_bucket[first_historical_date], 3)

    if any(lookup[product] for product in lookup):
        return lookup
    return fallback_historical_lookup(fallback_payload)


def fallback_historical_rows(payload: dict | None) -> list[dict]:
    payload = payload or {}
    existing = payload.get("historicalRows", [])
    return existing if isinstance(existing, list) else []


def build_historical_rows_from_rows(rows: list[dict], fallback_payload: dict | None = None) -> list[dict]:
    buckets: dict[tuple[str, str, str], dict] = {}
    for row in rows:
        product = row.get("product")
        if product not in {"MS", "HSD", "Power"}:
            continue
        date = str(row.get("date") or "")
        sap_code = normalize_code(row.get("sapCode"))
        if not date or not sap_code:
            continue
        historical_kl = parse_number(row.get("salesUnits"))
        if historical_kl <= 0:
            continue
        key = (date, sap_code, product)
        target = buckets.setdefault(
            key,
            {
                "date": date,
                "salesArea": row.get("salesArea", "UNKNOWN"),
                "salesAreaFull": row.get("salesAreaFull", row.get("salesArea", "UNKNOWN")),
                "sapCode": sap_code,
                "outletName": row.get("outletName", "UNKNOWN OUTLET"),
                "product": product,
                "historicalKl": 0.0,
                "plant": row.get("plant", ""),
                "isProjectAbhuyaday": bool(row.get("isProjectAbhuyaday", False)),
            },
        )
        target["historicalKl"] = round(target["historicalKl"] + historical_kl, 3)

    historical_rows = sorted(
        buckets.values(),
        key=lambda item: (
            item["date"],
            item["salesArea"],
            item["outletName"],
            item["product"],
        ),
    )
    return historical_rows or fallback_historical_rows(fallback_payload)


def build_historical_lookup(workbook, fallback_payload: dict | None = None, raw_rows: list[dict] | None = None) -> dict:
    if "Hist MS" not in workbook.sheetnames and "Hist HSD" not in workbook.sheetnames:
        if raw_rows:
            return build_historical_lookup_from_rows(raw_rows, fallback_payload)
        return fallback_historical_lookup(fallback_payload)

    lookup = fallback_historical_lookup(fallback_payload)
    for product, sheet_name in (("MS", "Hist MS"), ("HSD", "Hist HSD")):
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        headers = None
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if headers is None and values and len(values) > 1 and values[0] == "SAP CODE":
                headers = values
                continue
            if headers is None or not any(value is not None for value in values):
                continue
            sap_code = normalize_code(values[0])
            if not sap_code:
                continue
            for index, header in enumerate(headers[4:], start=4):
                month_day = normalize_month_day(header)
                if not month_day:
                    continue
                quantity = float(values[index] or 0)
                if quantity <= 0:
                    continue
                lookup[product][f"{sap_code}|{month_day}"] = round(quantity, 3)
    return lookup


def load_existing_payload(output_path: Path) -> dict:
    if not output_path.exists():
        return {}
    text = output_path.read_text(encoding="utf-8").strip()
    prefix = "window.SALES_DATA = "
    if not text.startswith(prefix):
        return {}
    json_text = text[len(prefix):]
    if json_text.endswith(";"):
        json_text = json_text[:-1]
    try:
        return json.loads(json_text)
    except json.JSONDecodeError:
        return {}


def build_payload(
    source_path: Path,
    fallback_payload: dict | None = None,
    historical_source_path: Path | None = None,
    outlet_master_path: Path | None = None,
    abhuyaday_path: Path | None = None,
) -> dict:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    abhuyaday_entries = load_abhuyaday_entries(abhuyaday_path)
    abhuyaday_codes = {entry["sapCode"] for entry in abhuyaday_entries}
    pa_codes = set(abhuyaday_codes) if abhuyaday_codes else read_pa_codes(workbook, fallback_payload)
    rows, product_meta = load_rows(workbook, pa_codes)
    outlet_master_entries = merge_abhuyaday_outlets(load_outlet_master_entries(outlet_master_path), abhuyaday_entries)
    rows = apply_outlet_master_to_rows(rows, outlet_master_entries)
    product_sequence = product_sequence_from_rows(rows)
    product_meta = build_product_meta(rows, product_sequence)
    historical_lookup = build_historical_lookup(workbook, fallback_payload)
    historical_meta = {
        "historicalSourceFile": source_path.name,
        "historicalDateMin": None,
        "historicalDateMax": None,
    }
    historical_rows_payload = fallback_historical_rows(fallback_payload)
    fallback_area_targets = (fallback_payload or {}).get("areaTargets", [])
    area_targets = build_abhuyaday_area_targets(abhuyaday_entries) or (
        fallback_area_targets if isinstance(fallback_area_targets, list) else []
    )
    abhuyaday_reference = build_abhuyaday_reference(abhuyaday_entries)

    if historical_source_path and historical_source_path.exists():
        historical_workbook = load_workbook(historical_source_path, read_only=True, data_only=True)
        try:
            historical_rows = load_raw_export_rows(historical_workbook, pa_codes)
            historical_rows = apply_outlet_master_to_rows(historical_rows, outlet_master_entries)
            historical_lookup = build_historical_lookup(historical_workbook, fallback_payload, historical_rows)
            historical_rows_payload = build_historical_rows_from_rows(historical_rows, fallback_payload)
            historical_dates = sorted({row["date"] for row in historical_rows_payload})
            historical_meta = {
                "historicalSourceFile": historical_source_path.name,
                "historicalDateMin": historical_dates[0] if historical_dates else None,
                "historicalDateMax": historical_dates[-1] if historical_dates else None,
            }
        finally:
            historical_workbook.close()
    workbook.close()

    all_dates = sorted({row["date"] for row in rows})
    all_areas = sorted({row["salesArea"] for row in rows})
    outlet_map = {}
    for entry in outlet_master_entries:
        outlet_map[entry["sapCode"]] = {
            "sapCode": entry["sapCode"],
            "outletName": entry["outletName"],
            "salesArea": entry["salesArea"],
            "isProjectAbhuyaday": entry["isProjectAbhuyaday"],
            "appointedOn": entry.get("appointedOn"),
            "plant": entry.get("plant", ""),
            "notes": entry.get("notes", ""),
            "outletType": entry.get("outletType", "Standard"),
        }

    for row in rows:
        outlet = outlet_map.setdefault(
            row["sapCode"],
            {
                "sapCode": row["sapCode"],
                "outletName": row["outletName"],
                "salesArea": row["salesArea"],
                "isProjectAbhuyaday": row["isProjectAbhuyaday"],
            },
        )
        if outlet["salesArea"] == "UNKNOWN" and row["salesArea"] != "UNKNOWN":
            outlet["salesArea"] = row["salesArea"]
        outlet["isProjectAbhuyaday"] = bool(outlet.get("isProjectAbhuyaday") or row.get("isProjectAbhuyaday"))

    date_counts = Counter(row["date"] for row in rows)

    return {
        "meta": {
            "title": "Jabalpur Retail Region Sales Dashboard",
            "sourceFile": source_path.name,
            "sourceMode": "Built-in workbook snapshot",
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "dateMin": all_dates[0] if all_dates else None,
            "dateMax": all_dates[-1] if all_dates else None,
            "defaultDate": all_dates[-1] if all_dates else None,
            "dateCount": len(all_dates),
            "areas": all_areas,
            "products": product_sequence,
            "emptyProducts": [name for name, meta in product_meta.items() if meta["records"] == 0],
            "paOutletCount": len(pa_codes),
            "projectAbhuyadayCodes": sorted(pa_codes),
            "projectAbhuyadaySourceFile": abhuyaday_path.name if abhuyaday_path and abhuyaday_path.exists() else None,
            "projectAbhuyadayReferenceCount": len(abhuyaday_entries),
            "activeOutletCount": len(outlet_map),
            "outletMasterSourceFile": outlet_master_path.name if outlet_master_path and outlet_master_path.exists() else None,
            "outletMasterCount": len(outlet_master_entries),
            "recordCount": len(rows),
            "productStats": product_meta,
            "recordsPerDate": dict(sorted(date_counts.items())),
            **historical_meta,
        },
        "filters": {
            "dates": all_dates,
            "areas": all_areas,
            "outlets": sorted(outlet_map.values(), key=lambda item: (item["salesArea"], item["outletName"])),
        },
        "historicalLookup": historical_lookup,
        "historicalRows": historical_rows_payload,
        "areaTargets": area_targets,
        "projectAbhuyadayReference": abhuyaday_reference,
        "rows": rows,
    }


def build_combined_payload(
    source_paths: list[Path],
    fallback_payload: dict | None = None,
    historical_source_path: Path | None = None,
    outlet_master_path: Path | None = None,
    abhuyaday_path: Path | None = None,
) -> dict:
    abhuyaday_entries = load_abhuyaday_entries(abhuyaday_path)
    abhuyaday_codes = {entry["sapCode"] for entry in abhuyaday_entries}
    outlet_master_entries = merge_abhuyaday_outlets(load_outlet_master_entries(outlet_master_path), abhuyaday_entries)
    rows: list[dict] = []
    pa_codes: set[str] = set()

    for source_path in source_paths:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            workbook_pa_codes = set(abhuyaday_codes) if abhuyaday_codes else read_pa_codes(workbook, fallback_payload)
            pa_codes.update(workbook_pa_codes)
            source_rows, _product_meta = load_rows(workbook, workbook_pa_codes)
            rows.extend(source_rows)
        finally:
            workbook.close()

    rows = apply_outlet_master_to_rows(rows, outlet_master_entries)
    rows = sorted(
        rows,
        key=lambda row: (
            row["date"],
            row["salesArea"],
            row["outletName"],
            row["product"],
            row.get("billingDocument", ""),
        ),
    )
    product_sequence = product_sequence_from_rows(rows)
    product_meta = build_product_meta(rows, product_sequence)
    historical_lookup = fallback_historical_lookup(fallback_payload)
    historical_meta = {
        "historicalSourceFile": None,
        "historicalDateMin": None,
        "historicalDateMax": None,
    }
    historical_rows_payload = fallback_historical_rows(fallback_payload)
    fallback_area_targets = (fallback_payload or {}).get("areaTargets", [])
    area_targets = build_abhuyaday_area_targets(abhuyaday_entries) or (
        fallback_area_targets if isinstance(fallback_area_targets, list) else []
    )
    abhuyaday_reference = build_abhuyaday_reference(abhuyaday_entries)

    if historical_source_path and historical_source_path.exists():
        historical_workbook = load_workbook(historical_source_path, read_only=True, data_only=True)
        try:
            historical_rows = load_raw_export_rows(historical_workbook, pa_codes)
            historical_rows = apply_outlet_master_to_rows(historical_rows, outlet_master_entries)
            historical_lookup = build_historical_lookup(historical_workbook, fallback_payload, historical_rows)
            historical_rows_payload = build_historical_rows_from_rows(historical_rows, fallback_payload)
            historical_dates = sorted({row["date"] for row in historical_rows_payload})
            historical_meta = {
                "historicalSourceFile": historical_source_path.name,
                "historicalDateMin": historical_dates[0] if historical_dates else None,
                "historicalDateMax": historical_dates[-1] if historical_dates else None,
            }
        finally:
            historical_workbook.close()

    all_dates = sorted({row["date"] for row in rows})
    all_areas = sorted({row["salesArea"] for row in rows})
    outlet_map = {}
    for entry in outlet_master_entries:
        outlet_map[entry["sapCode"]] = {
            "sapCode": entry["sapCode"],
            "outletName": entry["outletName"],
            "salesArea": entry["salesArea"],
            "isProjectAbhuyaday": entry["isProjectAbhuyaday"],
            "appointedOn": entry.get("appointedOn"),
            "plant": entry.get("plant", ""),
            "notes": entry.get("notes", ""),
            "outletType": entry.get("outletType", "Standard"),
        }

    for row in rows:
        outlet = outlet_map.setdefault(
            row["sapCode"],
            {
                "sapCode": row["sapCode"],
                "outletName": row["outletName"],
                "salesArea": row["salesArea"],
                "isProjectAbhuyaday": row["isProjectAbhuyaday"],
                "plant": row.get("plant", ""),
            },
        )
        if outlet["salesArea"] == "UNKNOWN" and row["salesArea"] != "UNKNOWN":
            outlet["salesArea"] = row["salesArea"]
        if not outlet.get("plant") and row.get("plant"):
            outlet["plant"] = row["plant"]
        outlet["isProjectAbhuyaday"] = bool(outlet.get("isProjectAbhuyaday") or row.get("isProjectAbhuyaday"))

    date_counts = Counter(row["date"] for row in rows)
    source_names = " + ".join(path.name for path in source_paths)

    return {
        "meta": {
            "title": "Jabalpur Retail Region Sales Dashboard",
            "sourceFile": source_names,
            "sourceMode": "Built-in workbook snapshot",
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "dateMin": all_dates[0] if all_dates else None,
            "dateMax": all_dates[-1] if all_dates else None,
            "defaultDate": all_dates[-1] if all_dates else None,
            "dateCount": len(all_dates),
            "areas": all_areas,
            "products": product_sequence,
            "emptyProducts": [name for name, meta in product_meta.items() if meta["records"] == 0],
            "paOutletCount": len(pa_codes),
            "projectAbhuyadayCodes": sorted(pa_codes),
            "projectAbhuyadaySourceFile": abhuyaday_path.name if abhuyaday_path and abhuyaday_path.exists() else None,
            "projectAbhuyadayReferenceCount": len(abhuyaday_entries),
            "activeOutletCount": len(outlet_map),
            "outletMasterSourceFile": outlet_master_path.name if outlet_master_path and outlet_master_path.exists() else None,
            "outletMasterCount": len(outlet_master_entries),
            "recordCount": len(rows),
            "productStats": product_meta,
            "recordsPerDate": dict(sorted(date_counts.items())),
            **historical_meta,
        },
        "filters": {
            "dates": all_dates,
            "areas": all_areas,
            "outlets": sorted(outlet_map.values(), key=lambda item: (item["salesArea"], item["outletName"])),
        },
        "historicalLookup": historical_lookup,
        "historicalRows": historical_rows_payload,
        "areaTargets": area_targets,
        "projectAbhuyadayReference": abhuyaday_reference,
        "rows": rows,
    }


def write_output(payload: dict, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    serialized = json.dumps(payload, separators=(",", ":"))
    output_path.write_text(f"window.SALES_DATA = {serialized};\n", encoding="utf-8")


def main() -> None:
    args = parse_args()
    source_paths = [Path(path) for path in args.input]
    output_path = Path(args.output)
    historical_source_path = Path(args.historical_input) if args.historical_input else None
    outlet_master_path = Path(args.outlet_master) if args.outlet_master else None
    abhuyaday_path = Path(args.abhuyaday_input) if args.abhuyaday_input else None
    fallback_payload = load_existing_payload(output_path)
    if len(source_paths) == 1:
        payload = build_payload(source_paths[0], fallback_payload, historical_source_path, outlet_master_path, abhuyaday_path)
    else:
        payload = build_combined_payload(source_paths, fallback_payload, historical_source_path, outlet_master_path, abhuyaday_path)
    write_output(payload, output_path)
    print(f"Generated {output_path}")


if __name__ == "__main__":
    main()
